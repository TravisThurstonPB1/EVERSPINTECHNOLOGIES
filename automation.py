import csv
from os import error
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Subledger
file_sub_ledget = 'Subleger.xls'

# FAB
file_FAB_promise = 'Promis_EWS_WIP.xls'

# MRAM LAB
file_MRAM_LAB = 'Promis_EWS_WIP.xls'

# Assembly
file_ASY_Amkor = 'Amkor_WIP_Report.xls'
file_CHM_Assy_WIP = 'wpsp006a_Everspin_ASSY.xls'
file_CHM_CP_WIP = 'CHM_CP_WIP.xls'
file_CHM_Assy_Inv = 'inventory_report.xls'
file_Daily_WIP_UDG = 'DAILY_WIP_DG.xls'
file_OSE_WIP = 'OSE WIP Report - EVERSPIN.xls'
file_UTC_Everspin_AssyWIP = 'EVERSPIN_AssyWIP.xls'
file_UTL_Assy = 'MSINV076.xls'

# UTC
file_UTC_WIP_report = 'UTC_EVERSPIN_WIP_REPORT.xls'

try:
	sub_ledger = pd.read_excel(file_sub_ledget, None)
	sub_ledger = pd.read_excel(open(file_sub_ledget, 'rb'),dtype={'Selection Date':datetime.date,'AsOfDate':datetime.date})
	sub_ledger = sub_ledger.replace(np.nan, '', regex=True)
except FileNotFoundError:
	print(file_sub_ledget + " file is required.")
	print("Automation process stopped!")
	exit(0)

try:
	FAB_promise = pd.read_excel(file_FAB_promise, None)
	FAB_promise = pd.read_excel(open(file_FAB_promise, 'rb'))
	FAB_promise = FAB_promise.replace(np.nan, '', regex=True)
except FileNotFoundError:
	print(file_FAB_promise + " file is required.")
	print("Automation process stopped!")
	exit(0)

try:
	MRAM_lab = pd.read_excel(file_MRAM_LAB, None)
	MRAM_lab = pd.read_excel(open(file_MRAM_LAB, 'rb'))
	MRAM_lab = MRAM_lab.replace(np.nan, '', regex=True)
except FileNotFoundError:
	print(file_MRAM_LAB + " file is required.")
	print("Automation process stopped!")
	exit(0)


### ASY WIP FILES ####
try:
	ASY_Amkor = pd.read_excel(file_ASY_Amkor, None)
	ASY_Amkor = pd.read_excel(open(file_ASY_Amkor, 'rb'))
	ASY_Amkor = ASY_Amkor.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

try:
	CHM_Assy_WIP = pd.read_excel(file_CHM_Assy_WIP, None)
	CHM_Assy_WIP = pd.read_excel(open(file_CHM_Assy_WIP, 'rb'))
	CHM_Assy_WIP = CHM_Assy_WIP.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass


try:
	CHM_CP_WIP = pd.read_excel(file_CHM_CP_WIP, None)
	CHM_CP_WIP = pd.read_excel(open(file_CHM_CP_WIP, 'rb'))
	CHM_CP_WIP = CHM_CP_WIP.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass
	# print(file_CHM_CP_WIP + " file is not found.")		

try:
	CHM_Assy_Inv = pd.read_excel(file_CHM_Assy_Inv, None)
	CHM_Assy_Inv = pd.read_excel(open(file_CHM_Assy_Inv, 'rb'))
	CHM_Assy_Inv = CHM_Assy_Inv.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

######################## IMPORTANT NOTE: HERE WE HAVE TO PASS SHEET NAME   ###################################### 
try:
	DAILY_WIP_UDG = pd.read_excel(file_Daily_WIP_UDG, None)
	DAILY_WIP_UDG = pd.read_excel(open(file_Daily_WIP_UDG, 'rb'),sheet_name="Detail")
	DAILY_WIP_UDG = DAILY_WIP_UDG.replace(np.nan, '', regex=True)
except FileNotFoundError:
	print(file_Daily_WIP_UDG + " file is not found.")
#################################################################################################################

try:
	OSE_WIP = pd.read_excel(file_OSE_WIP, None)
	OSE_WIP = pd.read_excel(open(file_OSE_WIP, 'rb'))
	OSE_WIP = OSE_WIP.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

try:
	UTC_EVERSPIN_AssyWIP = pd.read_excel(file_UTC_Everspin_AssyWIP, None)
	UTC_EVERSPIN_AssyWIP = pd.read_excel(open(file_UTC_Everspin_AssyWIP, 'rb'))
	UTC_EVERSPIN_AssyWIP = UTC_EVERSPIN_AssyWIP.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

try:
	UTL_Assy = pd.read_excel(file_UTL_Assy, None)
	UTL_Assy = pd.read_excel(open(file_UTL_Assy, 'rb'))
	UTL_Assy = UTL_Assy.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

######################### UTC TAB ###############################################
try:
	UTC_Tab_Sheet1 = pd.read_excel(file_UTC_WIP_report, None)
	UTC_Tab_Sheet1 = pd.read_excel(open(file_UTC_WIP_report, 'rb'))
	UTC_Tab_Sheet1 = UTC_Tab_Sheet1.replace(np.nan, '', regex=True)
except FileNotFoundError:
	pass

try:
	UTC_Tab_Sheet2 = pd.read_excel(file_UTC_WIP_report, None)
	if 'UTC Key' in UTC_Tab_Sheet2.keys():
		UTC_Tab_Sheet2 = pd.read_excel(open(file_UTC_WIP_report, 'rb'),sheet_name="UTC Key")
	else:
		try:
			UTC_Tab_Sheet2 = pd.read_excel(open('UTC Key.xlsx', 'rb'))
		except error:
			print("UTC Key.xlsx file is required to run.")
			exit(0)
except FileNotFoundError:
	try:
		UTC_Tab_Sheet2 = pd.read_excel(open('UTC Key.xlsx', 'rb'))
	except error:
		print("UTC Key.xlsx file is required to run.")
		exit(0)

UTC_Tab_Sheet2 = UTC_Tab_Sheet2.replace(np.nan, '', regex=True)

# Renaming the column name
dict = {'Opn_Code': 'OPN_CODE'} 
UTC_Tab_Sheet2.rename(columns=dict,inplace=True)
#################################################################################

wb = Workbook()
# Creating Worksheets
ws_sub_ledger_recon = wb.create_sheet("Subledger Recon", 0)
ws_FAB_pivot = wb.create_sheet("FAB Pivot", 1)
ws_FAB = wb.create_sheet("FAB", 2)
ws_MRAM_lab = wb.create_sheet("MRAM Lab", 3)
ws_MRAM_pivot = wb.create_sheet("MRAM Pivot", 4)
ws_asy_pivot = wb.create_sheet("Asy Pivot", 5)
ws_asy_WIP = wb.create_sheet("Asy WIP", 6)
ws_UTC_pivot = wb.create_sheet("UTC Pivot", 7)
ws_UTC = wb.create_sheet("UTC", 8)
ws_UTC_new = wb.create_sheet("UTC Pivot New", 9)
ws_subledger_UTC_pivot = wb.create_sheet("Subledger UTC Pivot", 10)

# Setting up Tabl Colors for wrorksheets
ws_sub_ledger_recon.sheet_properties.tabColor = "00B050"



for r in dataframe_to_rows(sub_ledger, index=False, header=True):
    ws_sub_ledger_recon.append(r)


# """
lot_number_array = []
ws_FAB_count = 1
for r in dataframe_to_rows(FAB_promise, index=False, header=True):
	if ws_FAB_count == 1:
		r.append("Lot Number")
		index_of_lot_comment = r.index("lotComment")
		index_of_lot_number = r.index("waferLot")
		ws_FAB.append(r)
		ws_FAB['Q1'] = "sub"
		ws_FAB['R1'] = "diff"
		ws_FAB_count = ws_FAB_count + 1
	if r[4] == "P" and r[8] != "":
		check = None
		# print(r[index_of_lot_comment])
		# exit(0)
		try:
			check = r[index_of_lot_comment].split()
			if len(check) > 1:
				r.append(r[index_of_lot_comment].split()[1])
				ws_FAB.append(r)
			else:
				r.append("")
				ws_FAB.append(r)
		except IndexError:
			pass
	if ws_FAB_count > 1 and r[4] == "P" and r[8] != "":
		if len(check) > 1:
			ws_FAB['Q'+str(ws_FAB_count)] = "=IFERROR(VLOOKUP(C"+str(ws_FAB_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(P"+str(ws_FAB_count)+",'Subledger Recon'!F:K,6,FALSE),"'#N/A'"))"
		else:
			ws_FAB['Q'+str(ws_FAB_count)] = "=IFERROR(VLOOKUP(C"+str(ws_FAB_count)+",'Subledger Recon'!F:K,6,FALSE),"'#N/A'")"
		ws_FAB['R'+str(ws_FAB_count)] = "=+K"+str(ws_FAB_count)+"-Q"+str(ws_FAB_count)+""
		ws_FAB_count = ws_FAB_count + 1
	if ws_FAB_count > 2 and r[4] == "P" and r[8] != "":
		check = r[index_of_lot_comment].split()
		if len(check) > 1:
			lot_number_array.append(r[index_of_lot_comment].split()[1])
		else:
			r = ""
			lot_number_array.append(r)
	
# print(len(lot_number_array))
# exit()
FAB_promise = FAB_promise[(FAB_promise['lotType'] == 'P') & (FAB_promise['currentStage'] != '')]
FAB_promise['Lot Number'] = lot_number_array
temp_df = FAB_promise[(FAB_promise['lotType'] == 'P')]
table = pd.pivot_table(temp_df, values=['currentQty'], index=['waferLot','Lot Number'],
                    aggfunc=np.sum, margins=True, margins_name='Total')
row = 2 # Start from this row number
col = 3 # Column 3 for Qty
col1 = 1 # Column 1 for Lot ID
col2 = 2 # Column 2 for Lot Number
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_FAB_pivot['A2'] = 'Lot Id'
	   ws_FAB_pivot['B2'] = 'Lot Number'
	   ws_FAB_pivot['C2'] = 'Qty'
	   try:
		   ws_FAB_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_FAB_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_FAB_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
	   except:
		   ws_FAB_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_FAB_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   pass

# """

ws_MRAM_lab_count = 1
for r in dataframe_to_rows(MRAM_lab, index=False, header=True):
	if ws_MRAM_lab_count == 1:
		ws_MRAM_lab.append(r)		
		ws_MRAM_lab['P1'] = "sub"
		ws_MRAM_lab['Q1'] = "diff"
		ws_MRAM_lab_count = ws_MRAM_lab_count + 1
	if r[4] == "P" and r[8] == "":
		ws_MRAM_lab.append(r)
	if ws_MRAM_lab_count > 1 and r[4] == "P" and r[8] == "":
		ws_MRAM_lab['P'+str(ws_MRAM_lab_count)] = "=IFERROR(VLOOKUP(C"+str(ws_MRAM_lab_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(D"+str(ws_MRAM_lab_count)+",'Subledger Recon'!F:K,6,FALSE),"'#N/A'"))"
		ws_MRAM_lab['Q'+str(ws_MRAM_lab_count)] = "=+K"+str(ws_MRAM_lab_count)+"-P"+str(ws_MRAM_lab_count)+""
		ws_MRAM_lab_count = ws_MRAM_lab_count + 1
	

temp_df = MRAM_lab[(MRAM_lab['lotType'] == 'P') & (MRAM_lab['currentStage'] == '')]
table = pd.pivot_table(temp_df, values=['currentQty'], index=['waferFamily','waferLot','promisLot'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

row = 2 # Start from this row number
col = 4 # Column 4 for Qty
col1 = 1 # Column 1 for waferFamily
col2 = 2 # Column 2 for waferLot
col3 = 3 # Column 3 for promisLot
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_MRAM_pivot['A2'] = 'WaferFamily'
	   ws_MRAM_pivot['B2'] = 'waferLot'
	   ws_MRAM_pivot['C2'] = 'promisLot'
	   ws_MRAM_pivot['D2'] = 'currentQty'
	   try:
		   ws_MRAM_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_MRAM_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_MRAM_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
		   ws_MRAM_pivot[get_column_letter(col3+c)+str(row+r)] = index[r][2]
	   except:
		   ws_MRAM_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_MRAM_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   pass


ws_sub_ledger_recon_count = 1
for r in dataframe_to_rows(sub_ledger, index=False, header=True):
	ws_sub_ledger_recon['R1'] = "Fab Tab"
	ws_sub_ledger_recon['R1'].fill = PatternFill("solid", fgColor="FFF2CC")
	ws_sub_ledger_recon['S1'] = "MRAM Tab"
	ws_sub_ledger_recon['S1'].fill = PatternFill("solid", fgColor="E2EFDA")
	ws_sub_ledger_recon['T1'] = "Assembly"
	ws_sub_ledger_recon['T1'].fill = PatternFill("solid", fgColor="D9E1F2")
	ws_sub_ledger_recon['U1'] = "Test"
	ws_sub_ledger_recon['U1'].fill = PatternFill("solid", fgColor="D0CECE")
	ws_sub_ledger_recon['V1'] = "Total"
	ws_sub_ledger_recon['W1'] = "Diff"
	ws_sub_ledger_recon['X1'] = "C Lot Sum"
	ws_sub_ledger_recon['Y1'] = "Lot No Sum"
	ws_sub_ledger_recon['Z1'] = "Parent Lot Sum"
	ws_sub_ledger_recon['AA1'] = "Source Lot Sum"
	ws_sub_ledger_recon['AB1'] = "C Lot"
	ws_sub_ledger_recon['AC1'] = "COUNTS"
	if ws_sub_ledger_recon_count > 1:
		ws_sub_ledger_recon['R'+str(ws_sub_ledger_recon_count)] = "=IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'FAB Pivot'!A:C,3,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'FAB Pivot'!B:C,2,FALSE),"'""'"))"
		ws_sub_ledger_recon['S'+str(ws_sub_ledger_recon_count)] = "=IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'MRAM Pivot'!B:D,3,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'MRAM Pivot'!C:D,2,FALSE),"'""'"))"
		ws_sub_ledger_recon['T'+str(ws_sub_ledger_recon_count)] = "=IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'Asy Pivot'!C:E,3,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'Asy Pivot'!D:E,2,FALSE),"'""'"))"
		ws_sub_ledger_recon['U'+str(ws_sub_ledger_recon_count)] = "=IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!A:B,2,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!D:F,3,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!I:J,2,FALSE),IFERROR(VLOOKUP(F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!L:M,2,FALSE),"'""'"))))"
		ws_sub_ledger_recon['V'+str(ws_sub_ledger_recon_count)] = "=SUM(R"+str(ws_sub_ledger_recon_count)+":U"+str(ws_sub_ledger_recon_count)+")"
		ws_sub_ledger_recon['W'+str(ws_sub_ledger_recon_count)] = "=+K"+str(ws_sub_ledger_recon_count)+"-V"+str(ws_sub_ledger_recon_count)
		ws_sub_ledger_recon['X'+str(ws_sub_ledger_recon_count)] = "=SUMIF('UTC Pivot'!A:A,'Subledger Recon'!F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!B:B)"
		ws_sub_ledger_recon['Y'+str(ws_sub_ledger_recon_count)] = "=SUMIF('UTC Pivot'!D:D,'Subledger Recon'!F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!F:F)"
		ws_sub_ledger_recon['Z'+str(ws_sub_ledger_recon_count)] = "=SUMIF('UTC Pivot'!I:I,'Subledger Recon'!F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!J:J)"
		ws_sub_ledger_recon['AA'+str(ws_sub_ledger_recon_count)] = "=SUMIF('UTC Pivot'!L:L,'Subledger Recon'!F"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!M:M)"
		ws_sub_ledger_recon['AB'+str(ws_sub_ledger_recon_count)] = "=VLOOKUP(G"+str(ws_sub_ledger_recon_count)+",'UTC Pivot'!A:B,2,FALSE)"
		ws_sub_ledger_recon['AC'+str(ws_sub_ledger_recon_count)] = "=COUNTIF('UTC Pivot'!H:H,'Subledger Recon'!G"+str(ws_sub_ledger_recon_count)+	")"

		 # Create fill
		r_column = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
		s_column = PatternFill(start_color='E2EFDA',end_color='E2EFDA',fill_type='solid')
		t_column = PatternFill(start_color='D9E1F2',end_color='D9E1F2',fill_type='solid')
		u_column = PatternFill(start_color='D0CECE',end_color='D0CECE',fill_type='solid')

		ws_sub_ledger_recon.conditional_formatting.add('W1:W100000',
             FormulaRule(formula=['R1=V1'], stopIfTrue=True, fill=r_column))
		ws_sub_ledger_recon.conditional_formatting.add('W1:W100000',
             FormulaRule(formula=['S1=V1'], stopIfTrue=True, fill=s_column))
		ws_sub_ledger_recon.conditional_formatting.add('W1:W100000',
             FormulaRule(formula=['T1=V1'], stopIfTrue=True, fill=t_column))
		ws_sub_ledger_recon.conditional_formatting.add('W1:W100000',
             FormulaRule(formula=['U1=V1'], stopIfTrue=True, fill=u_column))

	ws_sub_ledger_recon_count = ws_sub_ledger_recon_count + 1

ws_sub_ledger_recon['R'+str(ws_sub_ledger_recon_count+2)] = "FAB TAB"
ws_sub_ledger_recon['R'+str(ws_sub_ledger_recon_count+3)] = "=SUM(R2:R"+str(ws_sub_ledger_recon_count-1)+")"
ws_sub_ledger_recon['S'+str(ws_sub_ledger_recon_count+2)] = "MRAM TAB"
ws_sub_ledger_recon['S'+str(ws_sub_ledger_recon_count+3)] = "=SUM(S2:S"+str(ws_sub_ledger_recon_count-1)+")"
ws_sub_ledger_recon['T'+str(ws_sub_ledger_recon_count+2)] = "ASSEMBLY"
ws_sub_ledger_recon['T'+str(ws_sub_ledger_recon_count+3)] = "=SUM(T2:T"+str(ws_sub_ledger_recon_count-1)+")"
ws_sub_ledger_recon['U'+str(ws_sub_ledger_recon_count+2)] = "TEST"
ws_sub_ledger_recon['U'+str(ws_sub_ledger_recon_count+3)] = "=SUM(U2:U"+str(ws_sub_ledger_recon_count-1)+")"
ws_sub_ledger_recon['V'+str(ws_sub_ledger_recon_count+2)] = "TOTAL"
ws_sub_ledger_recon['V'+str(ws_sub_ledger_recon_count+3)] = "=SUM(V2:V"+str(ws_sub_ledger_recon_count-1)+")"
ws_sub_ledger_recon['W'+str(ws_sub_ledger_recon_count+2)] = "Diff"
ws_sub_ledger_recon['W'+str(ws_sub_ledger_recon_count+3)] = "=SUM(W2:W"+str(ws_sub_ledger_recon_count-1)+")"


############# ASY WIP ###################
ws_asy_WIP['A1'] = "PartID"
ws_asy_WIP['B1'] = "LotID"
ws_asy_WIP['C1'] = "ParentLot"
ws_asy_WIP['D1'] = "Qty"
ws_asy_WIP['E1'] = "Location"
ws_asy_WIP['F1'] = "Subledger Qty"
ws_asy_WIP['G1'] = "Diff Qty"


try:
	Amkor_ASY_count = 1
	for r in dataframe_to_rows(ASY_Amkor, index=False, header=True):
		if Amkor_ASY_count == 1 and r[0] == "Customer Number":
			Amkor_ASY_count = Amkor_ASY_count + 1
		elif Amkor_ASY_count > 1 and r[4] == "P3" and not(r[13].startswith("ES")) and r[13] != "":
			def checkQty(r):
				r = str(r)
				r = r.replace(" ", "")
				if r.isnumeric():
					return int(r)
				else:
					return r
			ws_asy_WIP['A'+str(Amkor_ASY_count)] = r[13] # PartID
			ws_asy_WIP['B'+str(Amkor_ASY_count)] = r[15] # LotID
			ws_asy_WIP['C'+str(Amkor_ASY_count)] = r[16] # ParentLot
			ws_asy_WIP['D'+str(Amkor_ASY_count)] = checkQty(r[62]) # Qty
			ws_asy_WIP['E'+str(Amkor_ASY_count)] = "Amkor" # Location
			ws_asy_WIP['F'+str(Amkor_ASY_count)] = "=IFERROR(VLOOKUP(B"+str(Amkor_ASY_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(Amkor_ASY_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(Amkor_ASY_count)] = "=D"+str(Amkor_ASY_count)+"-F"+str(Amkor_ASY_count)+""
			Amkor_ASY_count = Amkor_ASY_count + 1
except NameError:
	Amkor_ASY_count = 2
	pass

CHM_Assy_WIP_count = 0
try:
	for r in dataframe_to_rows(CHM_Assy_WIP, index=False, header=True):
		if CHM_Assy_WIP_count == 0 and r[0] == "Customer":
			CHM_Assy_WIP_count = Amkor_ASY_count
		elif CHM_Assy_WIP_count > 1 and r[5] != ""  and  not(r[5].startswith("ES")):
			if r[0] == "PACKAGE-ID":
				break
			def checkQty(r):
				row12 = str(r[12].replace(" ",""))
				row13 = str(r[13].replace(" ",""))
				row24 = str(r[24].replace(" ",""))

				if row24 and row24.isnumeric():
					return int(row24)
				elif row13 and row13.isnumeric():
					return int(row13)
				elif row12 and row12.isnumeric():
					return int(row12)
				else:
					total = 0
					# for index in range(14,24): # Taking column from O to X from xl file.
					# 	r[index] = r[index].replace(" ", "")
					# 	r[index] = str(r[index])
					# 	if r[index].isnumeric():
					# 		total = total + int(r[index])	
					return total
			
			ws_asy_WIP['A'+str(CHM_Assy_WIP_count)] = r[5] # PartID
			ws_asy_WIP['B'+str(CHM_Assy_WIP_count)] = r[6] # LotID
			ws_asy_WIP['C'+str(CHM_Assy_WIP_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(CHM_Assy_WIP_count)] = checkQty(r) # Qty
			ws_asy_WIP['E'+str(CHM_Assy_WIP_count)] = "CHM Assy WIP" # Location
			ws_asy_WIP['F'+str(CHM_Assy_WIP_count)] = "=IFERROR(VLOOKUP(B"+str(CHM_Assy_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(CHM_Assy_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(CHM_Assy_WIP_count)] = "=D"+str(CHM_Assy_WIP_count)+"-F"+str(CHM_Assy_WIP_count)+""
			CHM_Assy_WIP_count = CHM_Assy_WIP_count + 1
except NameError:
	CHM_Assy_WIP_count = Amkor_ASY_count
	pass	

try:
	CHM_CP_WIP_count = 0
	for r in dataframe_to_rows(CHM_CP_WIP, index=False, header=True):
		if CHM_CP_WIP_count == 0 and r[0] == "Priority":
			CHM_CP_WIP_count = CHM_Assy_WIP_count
		elif CHM_CP_WIP_count > 1 and r[13] != ""  and  not(r[13].startswith("Eng")):
			def checkQty(r):
				r = str(r)
				r = r.replace(" ", "")
				if r.isnumeric():
					return int(r)
				else:
					return r
			ws_asy_WIP['A'+str(CHM_CP_WIP_count)] = r[2] # PartID
			ws_asy_WIP['B'+str(CHM_CP_WIP_count)] = r[3] # LotID
			ws_asy_WIP['C'+str(CHM_CP_WIP_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(CHM_CP_WIP_count)] = checkQty(r[9]) # Qty
			ws_asy_WIP['E'+str(CHM_CP_WIP_count)] = "CHM_CP_WIP" # Location
			ws_asy_WIP['F'+str(CHM_CP_WIP_count)] = "=IFERROR(VLOOKUP(B"+str(CHM_CP_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(CHM_CP_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(CHM_CP_WIP_count)] = "=D"+str(CHM_CP_WIP_count)+"-F"+str(CHM_CP_WIP_count)+""
			CHM_CP_WIP_count = CHM_CP_WIP_count + 1
except NameError:
	CHM_CP_WIP_count = CHM_Assy_WIP_count
	pass

CHM_Assy_Inv_count = 0
try:
	for r in dataframe_to_rows(CHM_Assy_Inv, index=False, header=True):
		if CHM_Assy_Inv_count == 0 and r[0] == "PART_NO":
			CHM_Assy_Inv_count = CHM_CP_WIP_count
		elif CHM_Assy_Inv_count > 1 and r[0] != ""  and  not(r[0].startswith("ES")):
			def checkQty(r):
				r = str(r)
				r = r.replace(" ", "")
				if r.isnumeric():
					return int(r)
				else:
					return r
			ws_asy_WIP['A'+str(CHM_Assy_Inv_count)] = r[0] # PartID
			ws_asy_WIP['B'+str(CHM_Assy_Inv_count)] = r[2] # LotID
			ws_asy_WIP['C'+str(CHM_Assy_Inv_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(CHM_Assy_Inv_count)] = checkQty(r[5]) # Qty
			ws_asy_WIP['E'+str(CHM_Assy_Inv_count)] = "CHM_Assy_Inv" # Location
			ws_asy_WIP['F'+str(CHM_Assy_Inv_count)] = "=IFERROR(VLOOKUP(B"+str(CHM_Assy_Inv_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(CHM_Assy_Inv_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(CHM_Assy_Inv_count)] = "=D"+str(CHM_Assy_Inv_count)+"-F"+str(CHM_Assy_Inv_count)+""
			CHM_Assy_Inv_count = CHM_Assy_Inv_count + 1
except NameError:
	CHM_Assy_Inv_count = CHM_CP_WIP_count
	pass

try:
	DAILY_WIP_UDG_count = 0
	for r in dataframe_to_rows(DAILY_WIP_UDG, index=False, header=True):
		if DAILY_WIP_UDG_count == 0 and r[4] == "Customer Device Name":
			DAILY_WIP_UDG_count = CHM_Assy_Inv_count
		elif DAILY_WIP_UDG_count > 1 and r[4] != "":

			def checkLot(r):
				if len(r[8]) > 0:
					return r[8]
				else:
					return r[5]
			def checkQty(r):
				if len(r[13]) > 0:
					return int(r[13])
				else:
					total = 0
					for index in range(18,30): # Taking column from O to X from xl file.
						r[index] = str(r[index])
						r[index] = r[index].replace(" ", "")
						if r[index].isnumeric():
							total = total + int(r[index])	
					return total
				
			ws_asy_WIP['A'+str(DAILY_WIP_UDG_count)] = r[4] # PartID
			ws_asy_WIP['B'+str(DAILY_WIP_UDG_count)] = checkLot(r) # LotID
			ws_asy_WIP['C'+str(DAILY_WIP_UDG_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(DAILY_WIP_UDG_count)] = checkQty(r) # Qty
			ws_asy_WIP['E'+str(DAILY_WIP_UDG_count)] = "DAILY_WIP_UDG" # Location
			ws_asy_WIP['F'+str(DAILY_WIP_UDG_count)] = "=IFERROR(VLOOKUP(B"+str(DAILY_WIP_UDG_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(DAILY_WIP_UDG_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(DAILY_WIP_UDG_count)] = "=D"+str(DAILY_WIP_UDG_count)+"-F"+str(DAILY_WIP_UDG_count)+""
			DAILY_WIP_UDG_count = DAILY_WIP_UDG_count + 1
except NameError:
	DAILY_WIP_UDG_count = CHM_Assy_Inv_count
	pass

OSE_WIP_count = 0
try:
	for r in dataframe_to_rows(OSE_WIP, index=False, header=True):
		if OSE_WIP_count == 0 and r[1] == "DEVICE":
			OSE_WIP_count = DAILY_WIP_UDG_count
		elif OSE_WIP_count > 1 and r[1] != "" and not(r[1].startswith("ES")):
			def checkQty(r):
				total = 0
				for index in range(14,29): # Taking column from O to AC from xl file.
					r[index] = str(r[index])
					r[index] = r[index].replace(" ", "")
					if r[index].isnumeric():
						total = total + int(r[index])
				if total > 0:
					return total
				else:
					r[12] = str(r[12])
					if r[12].isnumeric() and int(r[12]) > 0:
						return int(r[12])
					else:
						return int(r[8])
			ws_asy_WIP['A'+str(OSE_WIP_count)] = r[1] # PartID
			ws_asy_WIP['B'+str(OSE_WIP_count)] = r[6] # LotID
			ws_asy_WIP['C'+str(OSE_WIP_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(OSE_WIP_count)] = checkQty(r) # Qty
			ws_asy_WIP['E'+str(OSE_WIP_count)] = "OSE_WIP" # Location
			ws_asy_WIP['F'+str(OSE_WIP_count)] = "=IFERROR(VLOOKUP(B"+str(OSE_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(OSE_WIP_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(OSE_WIP_count)] = "=D"+str(OSE_WIP_count)+"-F"+str(OSE_WIP_count)+""
			OSE_WIP_count = OSE_WIP_count + 1
except NameError:
	OSE_WIP_count = DAILY_WIP_UDG_count
	pass

UTC_EVERSPIN_AssyWIP_count = 0
try:
	for r in dataframe_to_rows(UTC_EVERSPIN_AssyWIP, index=False, header=True):
		if UTC_EVERSPIN_AssyWIP_count == 0 and r[4] == "DEVICE_PN":
			UTC_EVERSPIN_AssyWIP_count = OSE_WIP_count
		elif UTC_EVERSPIN_AssyWIP_count > 1 and r[4] != "":
			def checkQty(r):
				r = str(r)
				r = r.replace(" ", "")
				if r.isnumeric():
					return int(r)
				else:
					return r
			ws_asy_WIP['A'+str(UTC_EVERSPIN_AssyWIP_count)] = r[4] # PartID
			ws_asy_WIP['B'+str(UTC_EVERSPIN_AssyWIP_count)] = r[12] # LotID
			ws_asy_WIP['C'+str(UTC_EVERSPIN_AssyWIP_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(UTC_EVERSPIN_AssyWIP_count)] = checkQty(r[11]) # Qty
			ws_asy_WIP['E'+str(UTC_EVERSPIN_AssyWIP_count)] = "UTC_EVERSPIN_AssyWIP" # Location
			ws_asy_WIP['F'+str(UTC_EVERSPIN_AssyWIP_count)] = "=IFERROR(VLOOKUP(B"+str(UTC_EVERSPIN_AssyWIP_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(UTC_EVERSPIN_AssyWIP_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(UTC_EVERSPIN_AssyWIP_count)] = "=D"+str(UTC_EVERSPIN_AssyWIP_count)+"-F"+str(UTC_EVERSPIN_AssyWIP_count)+""
			UTC_EVERSPIN_AssyWIP_count = UTC_EVERSPIN_AssyWIP_count + 1
except NameError:
	UTC_EVERSPIN_AssyWIP_count = OSE_WIP_count
	pass

UTL_Assy_count = 0
try:
	for r in dataframe_to_rows(UTL_Assy, index=False, header=True):
		if UTL_Assy_count == 0 and r[1] == "Product No":
			UTL_Assy_count = UTC_EVERSPIN_AssyWIP_count
		elif UTL_Assy_count > 1 and r[1] != "" and not(r[1].startswith("ES")) and r[6] != "" and not(r[6].startswith("--")):
			def checkQty(r):
				r = str(r)
				r = r.replace(" ", "")
				if r.isnumeric():
					return int(r)
				else:
					return r
			ws_asy_WIP['A'+str(UTL_Assy_count)] = r[1] # PartID
			ws_asy_WIP['B'+str(UTL_Assy_count)] = r[6] # LotID
			ws_asy_WIP['C'+str(UTL_Assy_count)] = "" # ParentLot
			ws_asy_WIP['D'+str(UTL_Assy_count)] = checkQty(r[34]) # Qty
			ws_asy_WIP['E'+str(UTL_Assy_count)] = "UTL_Assy" # Location
			ws_asy_WIP['F'+str(UTL_Assy_count)] = "=IFERROR(VLOOKUP(B"+str(UTL_Assy_count)+",'Subledger Recon'!F:K,6,FALSE),IFERROR(VLOOKUP(C"+str(UTL_Assy_count)+",'Subledger Recon'!F:K,6,FALSE),""))"
			ws_asy_WIP['G'+str(UTL_Assy_count)] = "=D"+str(UTL_Assy_count)+"-F"+str(UTL_Assy_count)+""
			UTL_Assy_count = UTL_Assy_count + 1
except NameError:
	UTL_Assy_count = UTC_EVERSPIN_AssyWIP_count
	pass

ws_asy_WIP['A'+str(UTL_Assy_count)] = "Total"
ws_asy_WIP['D'+str(UTL_Assy_count)] = "=SUM(D2:D"+str(UTL_Assy_count-1)+")"
ws_asy_WIP['F'+str(UTL_Assy_count)] = "=SUM(F2:F"+str(UTL_Assy_count-1)+")"
ws_asy_WIP['G'+str(UTL_Assy_count)] = "=SUM(G2:G"+str(UTL_Assy_count-1)+")"

######### PIVOT TABLE FOR THE ASSEMBLY TAB ###################################
count = 1;partID= [];LotID= [];ParentLot= [];Qty= [];Location = [] 
df = pd.DataFrame(ws_asy_WIP.values)
for index, each in  df.iterrows():
	partID.append(each[0])
	LotID.append(each[1])
	ParentLot.append(each[2])
	Qty.append(each[3])
	Location.append(each[4])

# Removing Headers Name
partID.pop(0)
LotID.pop(0)
ParentLot.pop(0)
Qty.pop(0)
Location.pop(0)	

# initialize list of lists
data = {'PartID':partID, 'LotID':LotID, 'ParentLot':ParentLot, 'Qty': Qty, 'Location':Location}
assembly_df = pd.DataFrame(data)

table = pd.pivot_table(assembly_df, values=['Qty'], index=['Location','LotID'],
                    aggfunc=np.sum, margins=False, margins_name='Total')

# print(table)

row = 3 # Start from this row number
col = 3 # Column 3 total sum of Qty
col1 = 1 # Column 1 for Location
# col2 = 2 # Column 2 for PartID
col3 = 2 # Column 2 for LotID
# col4 = 4 # Column 2 for ParentLot
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_asy_pivot['A2'] = 'Location'
	   ws_asy_pivot['B2'] = 'LotID'
	   ws_asy_pivot['C2'] = 'Qty'
	#    ws_asy_pivot['D2'] = 'ParentLot'
	#    ws_asy_pivot['E2'] = 'Qty'
	   try:
		   ws_asy_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_asy_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		#    ws_asy_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
		   ws_asy_pivot[get_column_letter(col3+c)+str(row+r)] = index[r][1]
		#    ws_asy_pivot[get_column_letter(col4+c)+str(row+r)] = index[r][3]
	   except:
		   ws_asy_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_asy_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   pass


################################# UTC TABS ###############################################

UTC_main = pd.merge(UTC_Tab_Sheet1, 
                      UTC_Tab_Sheet2, 
                      on ='OPN_CODE', 
                      how ='inner')

count = 0
columns = ['LOT_NO', 'C_LOT', 'WAFER_LOT', 'OPN_CODE', 'STATUS', 'FLOW_CODE', 'DEVICE_NO', 'C_DEVICE', 'RECV_DATE', 'RECV_TIME', 'MOVE_IN_QTY', 'CHECK_IN_QTY', 'PART_NO', 'PACK_TYPE', 'PRE_FIRM', 'DATE_CODE', 'SOD', 'RESOD', 'SOURCE_LOT', 'LOT_KIND', 'PRODUCT_NAME', 'CUST_PONO', 'SO_NO', 'WO_NO', 'PARENT_LOT', 'MOVE_IN_DATE', 'MOVE_IN_TIME', 'ISSUED_QTY', 'Definition'] 
for r in dataframe_to_rows(UTC_main[columns], index=False, header=True):
	if count == 0:
		ws_UTC.append(r)
		count = count + 1
	elif count >= 1 and not(r[-1].startswith("ENG")) and not(r[-1].startswith("Eng")) and not(r[-1].startswith("Scrap")) and r[-1] != "":
		ws_UTC.append(r)
		count = count + 1

############ UTC PIVOT TABLE FOR C LOT #######################################
table = pd.pivot_table(UTC_main[columns], values=['MOVE_IN_QTY'], index=['C_LOT'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

row = 4 # Start from this row number
col = 2 # Column 3 total sum of Qty
col1 = 1 # Column 1 for C Lot number
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_UTC_pivot['A2'] = 'C_LOT'
	   ws_UTC_pivot['A3'] = 'C_LOT'
	   ws_UTC_pivot['B3'] = 'Sum of MOVE_IN_QTY'
	   try:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r]
	   except:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   pass


############ UTC PIVOT TABLE FOR LOT NO #######################################
table = pd.pivot_table(UTC_main[columns], values=['MOVE_IN_QTY'], index=['LOT_NO','C_LOT'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

row = 4 # Start from this row number
col = 6 # Column 3 total sum of Qty
col1 = 4 # Column 1 for Lot Number
col2 = 5 # Column 2 for C Lot number
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_UTC_pivot['D2'] = 'LOT_NO'
	   ws_UTC_pivot['D3'] = 'LOT_NO'
	   ws_UTC_pivot['E3'] = 'C_LOT'
	   ws_UTC_pivot['F3'] = 'Total'
	   try:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_UTC_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
	   except:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   pass


############ UTC PIVOT TABLE FOR PARENT LOT #######################################
table = pd.pivot_table(UTC_main[columns], values=['MOVE_IN_QTY'], index=['C_LOT','PARENT_LOT'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

row = 4 # Start from this row number
col = 10 # Column 3 total sum of Qty
col1 = 8 # Column 1 for Lot Number
col2 = 9 # Column 2 for C Lot number
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_UTC_pivot['H2'] = 'PARENT Lot'
	   ws_UTC_pivot['H3'] = 'C_LOT'
	   ws_UTC_pivot['I3'] = 'PARENT_LOT'
	   ws_UTC_pivot['J3'] = 'Total'
	   try:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_UTC_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
	   except:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   pass


############ UTC PIVOT TABLE FOR SOURCE LOT #######################################
table = pd.pivot_table(UTC_main[columns], values=['MOVE_IN_QTY'], index=['SOURCE_LOT'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

row = 4 # Start from this row number
col = 13 # Column 3 total sum of Qty
col1 = 12 # Column 1 for C Lot number
data = table.values
index = table.index
max_row, max_col = data.shape
for r in range(max_row):
   for c in range(max_col):
	   ws_UTC_pivot['L2'] = 'Source Lot'
	   ws_UTC_pivot['L3'] = 'SOURCE_LOT'
	   ws_UTC_pivot['M3'] = 'Sum of MOVE_IN_QTY'
	   try:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   ws_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r]
	   except:
		   ws_UTC_pivot[get_column_letter(col+c)+str(row+r)] = data[r][c]
		   pass


############ NEW UTC PIVOT TABLE FOR ALL LOTS NUMBER #######################################
table = pd.pivot_table(UTC_main[columns], values=['MOVE_IN_QTY'], index=['C_LOT','LOT_NO','PARENT_LOT','SOURCE_LOT'],
                    aggfunc=np.sum, margins=True, margins_name='Total')

# print(table)

row = 5 # Start from this row number
col1 = 1 # Column 1 for Lot Number
col2 = 2 # Column 2 for C Lot number
col3 = 3 # Column 2 for C Lot number
col4 = 4 # Column 2 for C Lot number
col5 = 5 # Column 3 total sum of Qty
data = table.values
index = table.index
max_row, max_col = data.shape
count = 4
for r in range(max_row):
   for c in range(max_col):
	   ws_UTC_new['A4'] = 'C_LOT'
	   ws_UTC_new['B4'] = 'LOT_NO'
	   ws_UTC_new['C4'] = 'PARENT_LOT'
	   ws_UTC_new['D4'] = 'SOURCE_LOT'
	   ws_UTC_new['E4'] = 'Total'
	   ws_UTC_new['F3'] = 'OnHand Lot'
	   ws_UTC_new['F4'] = 'LOT_NO'
	   ws_UTC_new['G3'] = 'OnHand ParentLot'
	   ws_UTC_new['G4'] = 'LOT_NO'
	   ws_UTC_new['H3'] = 'OnHand Lot'
	   ws_UTC_new['H4'] = 'C_LOT'
	   ws_UTC_new['I3'] = 'OnHand ParentLot'
	   ws_UTC_new['I4'] = 'C_LOT'

	   ws_UTC_new['J2'] = 'Remain Qty'
	   ws_UTC_new['K2'] = 'Remain Qty'
	   ws_UTC_new['L2'] = 'Remain Qty'
	   ws_UTC_new['M2'] = 'Remain Qty'

	   
	   ws_UTC_new['J3'] = 'OnHand Lot'
	   ws_UTC_new['K3'] = 'OnHand ParentLot'
	   ws_UTC_new['L3'] = 'OnHand Lot'
	   ws_UTC_new['M3'] = 'OnHand ParentLot'
	   ws_UTC_new['N3'] = 'Qty 3P'
	   ws_UTC_new['O3'] = 'Qty 3P'

	   ws_UTC_new['J4'] = 'LOT_NO'
	   ws_UTC_new['K4'] = 'LOT_NO'
	   ws_UTC_new['L4'] = 'C_LOT'
	   ws_UTC_new['M4'] = 'C_LOT'
	   ws_UTC_new['N4'] = 'Found'
	   ws_UTC_new['O4'] = 'Not Found'

	   try:
		   count = count + 1
		   ws_UTC_new[get_column_letter(col5+c)+str(row+r)] = data[r][c]
		   ws_UTC_new[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_UTC_new[get_column_letter(col2+c)+str(row+r)] = index[r][1]
		   ws_UTC_new[get_column_letter(col3+c)+str(row+r)] = index[r][2]
		   ws_UTC_new[get_column_letter(col4+c)+str(row+r)] = index[r][3]
		   ws_UTC_new[get_column_letter(6+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'Subledger Recon'!F:K,6,FALSE),0)"
		   ws_UTC_new[get_column_letter(7+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'Subledger Recon'!G:K,5,FALSE),0)"
		   ws_UTC_new[get_column_letter(8+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'Subledger Recon'!F:K,6,FALSE),0)"
		   ws_UTC_new[get_column_letter(9+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'Subledger Recon'!G:K,5,FALSE),0)"
		   ws_UTC_new[get_column_letter(10+c)+str(row+r)] = "=+E"+str(count)+"-F"+str(count)+""
		   ws_UTC_new[get_column_letter(11+c)+str(row+r)] = "=IF(J"+str(count)+"<>0,J"+str(count)+"-G"+str(count)+",J"+str(count)+")"
		   ws_UTC_new[get_column_letter(12+c)+str(row+r)] = "=IF(K"+str(count)+"<>0,K"+str(count)+"-H"+str(count)+",K"+str(count)+")"
		   ws_UTC_new[get_column_letter(13+c)+str(row+r)] = "=IF(L"+str(count)+"<>0,L"+str(count)+"-I"+str(count)+",L"+str(count)+")"
		   ws_UTC_new[get_column_letter(14+c)+str(row+r)] = "=IF(OR(F"+str(count)+">0,G"+str(count)+">0,H"+str(count)+">0,I"+str(count)+">0),E"+str(count)+","'""'")"
		   ws_UTC_new[get_column_letter(15+c)+str(row+r)] = "=IF(AND(F"+str(count)+"=0,G"+str(count)+"=0,H"+str(count)+"=0,I"+str(count)+"=0),E"+str(count)+","'""'")"
	   except:
		   count = count + 1
		   ws_UTC_new[get_column_letter(col5+c)+str(row+r)] = data[r][c]
		   ws_UTC_new[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_UTC_new[get_column_letter(col2+c)+str(row+r)] = index[r][1]
		   ws_UTC_new[get_column_letter(6+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'Subledger Recon'!F:K,6,FALSE),0)"
		   ws_UTC_new[get_column_letter(7+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'Subledger Recon'!G:K,5,FALSE),0)"
		   ws_UTC_new[get_column_letter(8+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'Subledger Recon'!F:K,6,FALSE),0)"
		   ws_UTC_new[get_column_letter(9+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'Subledger Recon'!G:K,5,FALSE),0)"
		   ws_UTC_new[get_column_letter(10+c)+str(row+r)] = "=+E"+str(count)+"-F"+str(count)+""
		   ws_UTC_new[get_column_letter(11+c)+str(row+r)] = "=IF(J"+str(count)+"<>0,J"+str(count)+"-G"+str(count)+",J"+str(count)+")"
		   ws_UTC_new[get_column_letter(12+c)+str(row+r)] = "=IF(K"+str(count)+"<>0,K"+str(count)+"-H"+str(count)+",K"+str(count)+")"
		   ws_UTC_new[get_column_letter(13+c)+str(row+r)] = "=IF(L"+str(count)+"<>0,L"+str(count)+"-I"+str(count)+",L"+str(count)+")"
		   ws_UTC_new[get_column_letter(14+c)+str(row+r)] = "=IF(OR(F"+str(count)+">0,G"+str(count)+">0,H"+str(count)+">0,I"+str(count)+">0),E"+str(count)+","'""'")"
		   ws_UTC_new[get_column_letter(15+c)+str(row+r)] = "=IF(AND(F"+str(count)+"=0,G"+str(count)+"=0,H"+str(count)+"=0,I"+str(count)+"=0),E"+str(count)+","'""'")"

ws_UTC_new['F'+str(count)] = "=SUM(F5:F"+str(count-1)+")"
ws_UTC_new['G'+str(count)] = "=SUM(G5:G"+str(count-1)+")"
ws_UTC_new['H'+str(count)] = "=SUM(H5:H"+str(count-1)+")"
ws_UTC_new['I'+str(count)] = "=SUM(I5:I"+str(count-1)+")"
ws_UTC_new['J'+str(count)] = "=SUM(J5:J"+str(count-1)+")"
ws_UTC_new['K'+str(count)] = "=SUM(K5:K"+str(count-1)+")"
ws_UTC_new['L'+str(count)] = "=SUM(L5:L"+str(count-1)+")"
ws_UTC_new['M'+str(count)] = "=SUM(M5:M"+str(count-1)+")"
ws_UTC_new['N'+str(count)] = "=SUM(N5:N"+str(count-1)+")"
ws_UTC_new['O'+str(count)] = "=SUM(O5:O"+str(count-1)+")"


# ws_subledger_UTC_pivot
# print(sub_ledger[['OnHand ParentLot','OnHand Lot','AbsEntry','OnHand/WIP Qty']])
# print(sub_ledger.columns.tolist())
############ NEW SUBLEDGER UTC PIVOT TABLE FOR OnHand Lot and OnHand Parent Lot #######################################
table = pd.pivot_table(sub_ledger[['OnHand ParentLot','OnHand Lot','AbsEntry','OnHand/WIP Qty']], values='OnHand/WIP Qty', index=['OnHand ParentLot','OnHand Lot','AbsEntry'],aggfunc=np.sum)
# print(table)


row = 4 # Start from this row number
col1 = 1 # Column 1 for OnHand ParentLot
col2 = 2 # Column 2 for OnHand Lot
col3 = 3 # Column 2 for AbsEntry
col4 = 4 # Column 2 for Total
data = table.values
index = table.index
max_row, max_col = data.shape
count = 4
for r in range(max_row):
   for c in range(max_col):
	   ws_subledger_UTC_pivot['A4'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['B4'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['C4'] = 'AbsEntry'
	   ws_subledger_UTC_pivot['D4'] = 'Total'

	   ws_subledger_UTC_pivot['E3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['F3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['G3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['H3'] = 'OnHand Lot'

	   ws_subledger_UTC_pivot['E4'] = 'LOT_NO'
	   ws_subledger_UTC_pivot['F4'] = 'C_LOT'
	   ws_subledger_UTC_pivot['G4'] = 'PARENT_LOT'
	   ws_subledger_UTC_pivot['H4'] = 'SOURCE_LOT'

	   ws_subledger_UTC_pivot['I3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['J3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['K3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['L3'] = 'OnHand ParentLot'

	   ws_subledger_UTC_pivot['I4'] = 'LOT_NO'
	   ws_subledger_UTC_pivot['J4'] = 'C_LOT'
	   ws_subledger_UTC_pivot['K4'] = 'PARENT_LOT'
	   ws_subledger_UTC_pivot['L4'] = 'SOURCE_LOT'
	   
	   ws_subledger_UTC_pivot['M2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['N2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['O2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['P2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['Q2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['R2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['S2'] = 'Remain Qty'
	   ws_subledger_UTC_pivot['T2'] = 'Remain Qty'

	   ws_subledger_UTC_pivot['M3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['N3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['O3'] = 'OnHand Lot'
	   ws_subledger_UTC_pivot['P3'] = 'OnHand Lot'

	   ws_subledger_UTC_pivot['M4'] = 'LOT_NO'
	   ws_subledger_UTC_pivot['N4'] = 'C_LOT'
	   ws_subledger_UTC_pivot['O4'] = 'PARENT_LOT'
	   ws_subledger_UTC_pivot['P4'] = 'SOURCE_LOT'

	   ws_subledger_UTC_pivot['Q3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['R3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['S3'] = 'OnHand ParentLot'
	   ws_subledger_UTC_pivot['T3'] = 'OnHand ParentLot'

	   ws_subledger_UTC_pivot['Q4'] = 'LOT_NO'
	   ws_subledger_UTC_pivot['R4'] = 'C_LOT'
	   ws_subledger_UTC_pivot['S4'] = 'PARENT_LOT'
	   ws_subledger_UTC_pivot['T4'] = 'SOURCE_LOT'
	   
	
	   ws_subledger_UTC_pivot['U3'] = 'Qty 3P'
	   ws_subledger_UTC_pivot['V3'] = 'Qty 3P'

	   ws_subledger_UTC_pivot['U4'] = 'Found'
	   ws_subledger_UTC_pivot['V4'] = 'Not Found'

	   try:
		   ws_subledger_UTC_pivot[get_column_letter(col4+c)+str(row+r)] = data[r][c]
		   ws_subledger_UTC_pivot[get_column_letter(col1+c)+str(row+r)] = index[r][0]
		   ws_subledger_UTC_pivot[get_column_letter(col2+c)+str(row+r)] = index[r][1]
		   ws_subledger_UTC_pivot[get_column_letter(col3+c)+str(row+r)] = index[r][2]
		   ws_subledger_UTC_pivot[get_column_letter(5+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'UTC Pivot New'!B:E,4,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(6+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'UTC Pivot New'!A:E,5,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(7+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'UTC Pivot New'!C:E,3,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(8+c)+str(row+r)] = "=IFERROR(VLOOKUP(B"+str(count)+",'UTC Pivot New'!D:E,2,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(9+c)+str(row+r)] = "==IFERROR(VLOOKUP(A"+str(count)+",'UTC Pivot New'!B:E,4,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(10+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'UTC Pivot New'!A:E,5,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(11+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'UTC Pivot New'!C:E,3,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(12+c)+str(row+r)] = "=IFERROR(VLOOKUP(A"+str(count)+",'UTC Pivot New'!D:E,2,FALSE),0)"
		   ws_subledger_UTC_pivot[get_column_letter(13+c)+str(row+r)] = "=+D"+str(count)+"-E"+str(count)+""
		   ws_subledger_UTC_pivot[get_column_letter(14+c)+str(row+r)] = "=IF(M"+str(count)+"<>0,M"+str(count)+"-F"+str(count)+",M"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(15+c)+str(row+r)] = "=IF(N"+str(count)+"<>0,N"+str(count)+"-G"+str(count)+",N"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(16+c)+str(row+r)] = "=IF(O"+str(count)+"<>0,O"+str(count)+"-H"+str(count)+",O"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(17+c)+str(row+r)] = "=IF(P"+str(count)+"<>0,P"+str(count)+"-I"+str(count)+",P"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(18+c)+str(row+r)] = "=IF(Q"+str(count)+"<>0,Q"+str(count)+"-J"+str(count)+",Q"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(19+c)+str(row+r)] = "=IF(R"+str(count)+"<>0,R"+str(count)+"-K"+str(count)+",R"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(20+c)+str(row+r)] = "=IF(S"+str(count)+"<>0,S"+str(count)+"-L"+str(count)+",S"+str(count)+")"
		   ws_subledger_UTC_pivot[get_column_letter(21+c)+str(row+r)] = "=IF(OR(E"+str(count)+">0,F"+str(count)+">0,G"+str(count)+">0,H"+str(count)+">0,I"+str(count)+">0,J"+str(count)+">0,K"+str(count)+">0,,L"+str(count)+">0),D"+str(count)+","'""'")"
		   ws_subledger_UTC_pivot[get_column_letter(22+c)+str(row+r)] = "=IF(AND(E"+str(count)+"=0,F"+str(count)+"=0,G"+str(count)+"=0,H"+str(count)+"=0,I"+str(count)+"=0,J"+str(count)+"=0,K"+str(count)+"=0,L"+str(count)+"=0),D"+str(count)+","'""'")"
		   count = count + 1
	   except:
		   print('exception')
		   exit(0)
		   pass

# ws_subledger_UTC_pivot['F'+str(count)] = "=SUM(F5:F"+str(count-1)+")"
# ws_subledger_UTC_pivot['G'+str(count)] = "=SUM(G5:G"+str(count-1)+")"
# ws_subledger_UTC_pivot['H'+str(count)] = "=SUM(H5:H"+str(count-1)+")"
# ws_subledger_UTC_pivot['I'+str(count)] = "=SUM(I5:I"+str(count-1)+")"
# ws_subledger_UTC_pivot['J'+str(count)] = "=SUM(J5:J"+str(count-1)+")"
# ws_subledger_UTC_pivot['K'+str(count)] = "=SUM(K5:K"+str(count-1)+")"
# ws_subledger_UTC_pivot['L'+str(count)] = "=SUM(L5:L"+str(count-1)+")"
# ws_subledger_UTC_pivot['M'+str(count)] = "=SUM(M5:M"+str(count-1)+")"
# ws_subledger_UTC_pivot['N'+str(count)] = "=SUM(N5:N"+str(count-1)+")"
# ws_subledger_UTC_pivot['O'+str(count)] = "=SUM(O5:O"+str(count-1)+")"









import datetime
x = datetime.datetime.now()
if x.month < 10:
    month = '0'+str(x.month)
else:
    month = x.month

wb.save('output - '+ str(x.day)+'-'+ str(month)+'-'+ str(x.year) +'.xlsx')
# wb.save('output - '+ str(30)+'-'+ str('09')+'-'+ str(x.year) +'.xlsx')
# print('output - '+ str(30)+'-'+ str('09')+'-'+ str(x.year) +'.xlsx')
print("Process Done!")





